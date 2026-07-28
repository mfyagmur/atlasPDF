import io
import zipfile

import pikepdf
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfWriter

from app.main import app

client = TestClient(app)


def make_blank_pdf_bytes(num_pages: int = 1) -> bytes:
    writer = PdfWriter()
    for _ in range(num_pages):
        writer.add_blank_page(width=200, height=200)
    buffer = io.BytesIO()
    writer.write(buffer)
    return buffer.getvalue()


def make_table_pdf_bytes(
    tables: list[list[list[str]]],
    cell_w: int = 100,
    cell_h: int = 30,
    origin: tuple[int, int] = (50, 300),
) -> bytes:
    pdf = pikepdf.Pdf.new()
    x0, y0 = origin

    for table in tables:
        rows = len(table)
        cols = len(table[0])
        page = pdf.add_blank_page(page_size=(50 + cols * cell_w + 50, 400))

        ops = []
        for r in range(rows + 1):
            y = y0 - r * cell_h
            ops.append(f"{x0} {y} m {x0 + cols * cell_w} {y} l S")
        for c in range(cols + 1):
            x = x0 + c * cell_w
            ops.append(f"{x} {y0} m {x} {y0 - rows * cell_h} l S")

        text_ops = ["BT /F1 12 Tf"]
        for r in range(rows):
            for c in range(cols):
                tx = x0 + c * cell_w + 5
                ty = y0 - r * cell_h - cell_h + 10
                text_ops.append(f"1 0 0 1 {tx} {ty} Tm ({table[r][c]}) Tj")
        text_ops.append("ET")

        content = ("\n".join(ops) + "\n" + "\n".join(text_ops)).encode()
        stream = pdf.make_stream(content)
        font = pikepdf.Dictionary(
            Type=pikepdf.Name.Font,
            Subtype=pikepdf.Name.Type1,
            BaseFont=pikepdf.Name.Helvetica,
        )
        page.obj.Resources = pikepdf.Dictionary(Font=pikepdf.Dictionary(F1=font))
        page.obj.Contents = stream

    buffer = io.BytesIO()
    pdf.save(buffer)
    return buffer.getvalue()


def make_encrypted_pdf_bytes() -> bytes:
    writer = PdfWriter()
    writer.add_blank_page(width=200, height=200)
    writer.encrypt("secret")
    buffer = io.BytesIO()
    writer.write(buffer)
    return buffer.getvalue()


def _add_page_with_content(
    pdf: "pikepdf.Pdf",
    table: list[list[str]] | None,
    extra_text: str | None,
    cell_w: int = 100,
    cell_h: int = 30,
    origin: tuple[int, int] = (50, 300),
) -> None:
    x0, y0 = origin
    ops: list[str] = []
    text_ops = ["BT /F1 12 Tf"]

    if table is not None:
        rows = len(table)
        cols = len(table[0])
        page = pdf.add_blank_page(page_size=(50 + cols * cell_w + 50, 400))

        for r in range(rows + 1):
            y = y0 - r * cell_h
            ops.append(f"{x0} {y} m {x0 + cols * cell_w} {y} l S")
        for c in range(cols + 1):
            x = x0 + c * cell_w
            ops.append(f"{x} {y0} m {x} {y0 - rows * cell_h} l S")

        for r in range(rows):
            for c in range(cols):
                tx = x0 + c * cell_w + 5
                ty = y0 - r * cell_h - cell_h + 10
                text_ops.append(f"1 0 0 1 {tx} {ty} Tm ({table[r][c]}) Tj")
    else:
        page = pdf.add_blank_page(page_size=(400, 400))

    if extra_text is not None:
        text_ops.append(f"1 0 0 1 {x0} {y0 - 350} Tm ({extra_text}) Tj")

    text_ops.append("ET")

    content = ("\n".join(ops) + "\n" + "\n".join(text_ops)).encode()
    stream = pdf.make_stream(content)
    font = pikepdf.Dictionary(
        Type=pikepdf.Name.Font,
        Subtype=pikepdf.Name.Type1,
        BaseFont=pikepdf.Name.Helvetica,
    )
    page.obj.Resources = pikepdf.Dictionary(Font=pikepdf.Dictionary(F1=font))
    page.obj.Contents = stream


def make_watermarked_table_pdf_bytes() -> bytes:
    table = [["a", "b", "c"], ["d", "e", "f"], ["g", "h", "i"]]
    cell_w, cell_h = 100, 30
    x0, y0 = 50, 300
    rows, cols = len(table), len(table[0])

    pdf = pikepdf.Pdf.new()
    page = pdf.add_blank_page(page_size=(50 + cols * cell_w + 50, 400))

    ops = []
    for r in range(rows + 1):
        y = y0 - r * cell_h
        ops.append(f"{x0} {y} m {x0 + cols * cell_w} {y} l S")
    for c in range(cols + 1):
        x = x0 + c * cell_w
        ops.append(f"{x} {y0} m {x} {y0 - rows * cell_h} l S")

    text_ops = [
        "0.85 g",
        "BT /F1 40 Tf",
        "0.7071 0.7071 -0.7071 0.7071 60 150 Tm (FILIGRAN ORNEK METNI COK UZUN BIR YAZI) Tj",
        "ET",
        "0 g",
        "BT /F1 12 Tf",
    ]
    for r in range(rows):
        for c in range(cols):
            tx = x0 + c * cell_w + 5
            ty = y0 - r * cell_h - cell_h + 10
            text_ops.append(f"1 0 0 1 {tx} {ty} Tm ({table[r][c]}) Tj")
    text_ops.append("ET")

    content = ("\n".join(ops) + "\n" + "\n".join(text_ops)).encode()
    stream = pdf.make_stream(content)
    font = pikepdf.Dictionary(
        Type=pikepdf.Name.Font,
        Subtype=pikepdf.Name.Type1,
        BaseFont=pikepdf.Name.Helvetica,
    )
    page.obj.Resources = pikepdf.Dictionary(Font=pikepdf.Dictionary(F1=font))
    page.obj.Contents = stream

    buffer = io.BytesIO()
    pdf.save(buffer)
    return buffer.getvalue()


def make_mixed_pdf_bytes() -> bytes:
    pdf = pikepdf.Pdf.new()
    _add_page_with_content(pdf, table=[["a", "b"], ["c", "d"]], extra_text=None)
    _add_page_with_content(pdf, table=None, extra_text="sadece metin")
    _add_page_with_content(pdf, table=[["x", "y"]], extra_text="tablo altinda metin")

    buffer = io.BytesIO()
    pdf.save(buffer)
    return buffer.getvalue()


def test_pdf_to_excel_happy_path():
    table = [["a", "b", "c"], ["d", "e", "f"]]
    pdf = make_table_pdf_bytes([table])

    response = client.post(
        "/api/pdf-to-excel",
        files={"file": ("a.pdf", pdf, "application/pdf")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["tables_found"] == 1
    assert data["warning"] is None
    assert data["download_url"] == f"/api/download/{data['file_id']}"

    download = client.get(data["download_url"])
    assert download.status_code == 200
    assert download.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert zipfile.is_zipfile(io.BytesIO(download.content))

    workbook = load_workbook(io.BytesIO(download.content))
    sheet = workbook["Sayfa1"]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    assert rows == table


def test_pdf_to_excel_multiple_tables():
    table_1 = [["a", "b"], ["c", "d"]]
    table_2 = [["x", "y"], ["z", "w"]]
    pdf = make_table_pdf_bytes([table_1, table_2])

    response = client.post(
        "/api/pdf-to-excel",
        files={"file": ("multi.pdf", pdf, "application/pdf")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["tables_found"] == 2

    download = client.get(data["download_url"])
    workbook = load_workbook(io.BytesIO(download.content))
    assert workbook.sheetnames == ["Sayfa1", "Sayfa2"]


def test_pdf_to_excel_no_table_fallback():
    pdf = make_blank_pdf_bytes(2)

    response = client.post(
        "/api/pdf-to-excel",
        files={"file": ("blank.pdf", pdf, "application/pdf")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["tables_found"] == 0
    assert data["warning"] is not None

    download = client.get(data["download_url"])
    assert download.status_code == 200
    workbook = load_workbook(io.BytesIO(download.content))
    assert workbook.sheetnames == ["Sayfa1", "Sayfa2"]


def test_pdf_to_excel_rejects_non_pdf_file():
    fake = b"hello world, this is not a pdf"

    response = client.post(
        "/api/pdf-to-excel",
        files={"file": ("fake.pdf", fake, "application/pdf")},
    )

    assert response.status_code == 415


def test_pdf_to_excel_rejects_encrypted_pdf():
    pdf = make_encrypted_pdf_bytes()

    response = client.post(
        "/api/pdf-to-excel",
        files={"file": ("encrypted.pdf", pdf, "application/pdf")},
    )

    assert response.status_code == 422
    assert "Şifreli" in response.json()["detail"]


def test_pdf_to_excel_ignores_watermark_text_in_table():
    pdf = make_watermarked_table_pdf_bytes()

    response = client.post(
        "/api/pdf-to-excel",
        files={"file": ("watermark.pdf", pdf, "application/pdf")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["tables_found"] == 1

    download = client.get(data["download_url"])
    workbook = load_workbook(io.BytesIO(download.content))
    sheet = workbook["Sayfa1"]

    all_cells = [cell.value for row in sheet.iter_rows() for cell in row]
    assert all(v is None or len(v) < 100 for v in all_cells)
    assert not any(v and "FILIGRAN" in v for v in all_cells)

    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    assert rows == [["a", "b", "c"], ["d", "e", "f"], ["g", "h", "i"]]


def test_pdf_to_excel_page_count_matches_sheet_count():
    pdf = make_mixed_pdf_bytes()

    response = client.post(
        "/api/pdf-to-excel",
        files={"file": ("mixed.pdf", pdf, "application/pdf")},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["tables_found"] == 2

    download = client.get(data["download_url"])
    workbook = load_workbook(io.BytesIO(download.content))
    assert workbook.sheetnames == ["Sayfa1", "Sayfa2", "Sayfa3"]

    sheet2_values = [cell.value for row in workbook["Sayfa2"].iter_rows() for cell in row]
    assert any(v and "sadece metin" in v for v in sheet2_values)

    sheet3_rows = [[cell.value for cell in row] for row in workbook["Sayfa3"].iter_rows()]
    assert ["x", "y"] in sheet3_rows
    flat_sheet3 = [v for row in sheet3_rows for v in row]
    assert any(v and "tablo altinda metin" in v for v in flat_sheet3)
